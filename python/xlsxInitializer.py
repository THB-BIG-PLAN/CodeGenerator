import os
import warnings
from openpyxl.styles import Font, Alignment
import pandas as pd
import openpyxl
import re

warnings.simplefilter(action='ignore', category=UserWarning)
HAND_CONFIG_PATH = "ConfigByHand.xlsx"
InputSignal_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="InputSignal")
CONFIG_PATH = "Config.xlsx"
List_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="List")
Condition_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="Condition")
EVT_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="EVT")
TimeFlag_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="TimeFlag")
ConstMacro_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="ConstMacro")
OutputInitializer_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="OutputInitializer")
Action_DataFrame = pd.read_excel(HAND_CONFIG_PATH, sheet_name="Action")
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
order = ['InputSignal', 'OutputInitializer', 'Condition', 'EVT', 'TimeFlag', 'ConstMacro', 'Action', 'List']


def WriteInputSignalToExcel():
    NewSignal_DataFrame = InputSignal_DataFrame.copy()
    new_rows = []
    for i, row in InputSignal_DataFrame.iterrows():
        SignalPre_Name = row.loc['SignalName'] + "_Pre"
        new_rows.append({'SignalName': SignalPre_Name, 'Type': row.loc['Type']})
    NewSignal_DataFrame = pd.concat([NewSignal_DataFrame, pd.DataFrame(new_rows)], ignore_index=True)
    NewSignal_DataFrame['Macro'] = NewSignal_DataFrame['SignalName'] + '_SIGNALNUM'
    # order by Type
    NewSignal_DataFrame = NewSignal_DataFrame.sort_values(by=['Type', 'SignalName'])
    NewSignal_DataFrame.reset_index(drop=True, inplace=True)
    # write to excel
    if not os.path.exists(CONFIG_PATH):
        NewSignal_DataFrame.to_excel(CONFIG_PATH, sheet_name='InputSignal', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            NewSignal_DataFrame.to_excel(writer, sheet_name='InputSignal', index=False)


def WriteConditionToExcel():
    NewRowList = []
    NewRow = {}
    NewCondition_DataFrame = pd.DataFrame(
        columns=['SignalName', 'Symbol', 'Threshold', 'ThresholdType', 'EVT', 'Macro'])
    for i, row in Condition_DataFrame.iterrows():
        Signal_Name = row.loc['SignalName']
        if row.loc['Symbol'] == 'CHANGE':
            pattern = r'\((.*),(.*)\)'
            match = re.match(pattern, row.loc['Threshold'])
            a = match.group(1)
            b = match.group(2)
            if a == 'x' and b == 'x':
                NewRow = {
                    'SignalName': Signal_Name,
                    'Symbol': row.loc['Symbol'],
                    'Threshold': Signal_Name + '_Pre_SIGNALNUM',
                    'ThresholdType': 'CONDITION_TYPE_SIGNAL',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_CHANGE_' + Signal_Name.upper() + '_PRE'
                }
                NewRowList.append(NewRow)
            elif a != 'x' and b == 'x':
                NewRow = {
                    'SignalName': Signal_Name,
                    'Symbol': row.loc['Symbol'],
                    'Threshold': Signal_Name + '_Pre_SIGNALNUM',
                    'ThresholdType': 'CONDITION_TYPE_SIGNAL',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_CHANGE_' + Signal_Name.upper() + '_PRE'
                }
                NewRowList.append(NewRow)
                NewRow = {
                    'SignalName': Signal_Name + '_Pre',
                    'Symbol': 'EQ',
                    'Threshold': a,
                    'ThresholdType': 'CONDITION_TYPE_NUMBER',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_PRE_EQ_' + a
                }
                NewRowList.append(NewRow)
            elif a == 'x' and b != 'x':
                NewRow = {
                    'SignalName': Signal_Name,
                    'Symbol': row.loc['Symbol'],
                    'Threshold': Signal_Name + '_Pre_SIGNALNUM',
                    'ThresholdType': 'CONDITION_TYPE_SIGNAL',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_CHANGE_' + Signal_Name.upper() + '_PRE'
                }
                NewRowList.append(NewRow)
                NewRow = {
                    'SignalName': Signal_Name,
                    'Symbol': 'EQ',
                    'Threshold': b,
                    'ThresholdType': 'CONDITION_TYPE_NUMBER',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_EQ_' + b
                }
                NewRowList.append(NewRow)
            elif a != 'x' and b != 'x':
                NewRow = {
                    'SignalName': Signal_Name + '_Pre',
                    'Symbol': 'EQ',
                    'Threshold': a,
                    'ThresholdType': 'CONDITION_TYPE_NUMBER',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_PRE_EQ_' + a
                }
                NewRowList.append(NewRow)
                NewRow = {
                    'SignalName': Signal_Name,
                    'Symbol': 'EQ',
                    'Threshold': b,
                    'ThresholdType': 'CONDITION_TYPE_NUMBER',
                    'EVT': row.loc['EVT'],
                    'Macro': Signal_Name.upper() + '_EQ_' + b
                }
                NewRowList.append(NewRow)
        else:
            Threshold = row.loc['Threshold']
            NewRow = {
                'SignalName': row.loc['SignalName'],
                'Symbol': row.loc['Symbol'],
                'EVT': row.loc['EVT'],
                'Macro': row.loc['Macro']
            }
            if InputSignal_DataFrame['SignalName'].isin([Threshold]).any():
                NewRow['ThresholdType'] = 'CONDITION_TYPE_SIGNAL'
                NewRow['Threshold'] = Threshold + '_SIGNALNUM'
            else:
                NewRow['ThresholdType'] = 'CONDITION_TYPE_NUMBER'
                NewRow['Threshold'] = Threshold
            NewRowList.append(NewRow)
    NewCondition_DataFrame = pd.concat([NewCondition_DataFrame, pd.DataFrame(NewRowList)], ignore_index=True)
    # order by SignalName
    NewCondition_DataFrame = NewCondition_DataFrame.sort_values(by=['SignalName'])
    NewCondition_DataFrame.reset_index(drop=True, inplace=True)
    # write to excel
    if not os.path.exists(CONFIG_PATH):
        NewCondition_DataFrame.to_excel(CONFIG_PATH, sheet_name='Condition', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            NewCondition_DataFrame.to_excel(writer, sheet_name='Condition', index=False)


def WriteEVTToExcel():
    NewRowList = []
    NewRow = {}
    NewEVT_DataFrame = pd.DataFrame(
        columns=['EVTName', 'Action1', 'Action2', 'Condition0', 'Condition1',
                 'Condition2', 'Condition3', 'Condition4', 'Condition5', 'Condition6', 'Condition7', 'Condition8'])
    for i, row in EVT_DataFrame.iterrows():
        NewRow = {}
        ConditionNum = 0
        for j, col in enumerate(row):
            if j <= 2:
                NewRow[row.index[j]] = col
                continue
            if pd.notna(col) and col != 'AND':
                col = str(col)
                pattern = r'^.*CHANGE.*\(([^)]+),([^)]+)\).*$'
                match = re.match(pattern, col)
                if match:
                    a = match.group(1)
                    b = match.group(2)
                    matching_rows = Condition_DataFrame[Condition_DataFrame['Macro'] == col]
                    # print(matching_rows['SignalName'])
                    SignalString = str(matching_rows['SignalName'].iloc[0]).upper()
                    if a == 'X' and b == 'X':
                        NewRow['Condition' + str(
                            ConditionNum)] = SignalString + '_CHANGE_' + SignalString + '_PRE'
                        ConditionNum += 1
                        pass
                    elif a != 'X' and b == 'X':
                        NewRow['Condition' + str(
                            ConditionNum)] = SignalString + '_CHANGE_' + SignalString + '_PRE'
                        ConditionNum += 1
                        NewRow['Condition' + str(ConditionNum)] = SignalString + '_PRE_EQ_' + a
                        ConditionNum += 1
                        pass
                    elif a == 'X' and b != 'X':
                        NewRow['Condition' + str(
                            ConditionNum)] = SignalString + '_CHANGE_' + SignalString + '_PRE'
                        ConditionNum += 1
                        NewRow['Condition' + str(ConditionNum)] = SignalString + '_EQ_' + b
                        ConditionNum += 1
                        pass
                    elif a != 'X' and b != 'X':
                        NewRow['Condition' + str(ConditionNum)] = SignalString + '_EQ_' + b
                        ConditionNum += 1
                        NewRow['Condition' + str(ConditionNum)] = SignalString + '_PRE_EQ_' + a
                        ConditionNum += 1
                        pass
                else:
                    NewRow['Condition' + str(ConditionNum)] = col
                    ConditionNum += 1
        NewRowList.append(NewRow)
    NewEVT_DataFrame = pd.concat([NewEVT_DataFrame, pd.DataFrame(NewRowList)], ignore_index=True)
    # write to excel
    if not os.path.exists(CONFIG_PATH):
        NewEVT_DataFrame.to_excel(CONFIG_PATH, sheet_name='EVT', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            NewEVT_DataFrame.to_excel(writer, sheet_name='EVT', index=False)


def WriteTimeFlagToExcel():
    if not os.path.exists(CONFIG_PATH):
        TimeFlag_DataFrame.to_excel(CONFIG_PATH, sheet_name='TimeFlag', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            TimeFlag_DataFrame.to_excel(writer, sheet_name='TimeFlag', index=False)


def WriteConstMacroToExcel():
    if not os.path.exists(CONFIG_PATH):
        ConstMacro_DataFrame.to_excel(CONFIG_PATH, sheet_name='ConstMacro', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            ConstMacro_DataFrame.to_excel(writer, sheet_name='ConstMacro', index=False)


def WriteOutputInitializerToExcel():
    if not os.path.exists(CONFIG_PATH):
        OutputInitializer_DataFrame.to_excel(CONFIG_PATH, sheet_name='OutputInitializer', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            OutputInitializer_DataFrame.to_excel(writer, sheet_name='OutputInitializer', index=False)


def WriteActionToExcel():
    if not os.path.exists(CONFIG_PATH):
        Action_DataFrame.to_excel(CONFIG_PATH, sheet_name='Action', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            Action_DataFrame.to_excel(writer, sheet_name='Action', index=False)


def WriteListToExcel():
    NewCondition_DataFrame = pd.read_excel(CONFIG_PATH, sheet_name="Condition")
    condition_macro_unique = NewCondition_DataFrame['Macro'].drop_duplicates().reset_index(drop=True)
    new_row = pd.Series(['TIMEFLAGNUM_EQ_0'])
    condition_macro_unique = pd.concat([condition_macro_unique, new_row], ignore_index=True)
    NewList_DataFrame = List_DataFrame.copy()
    if len(condition_macro_unique) > len(NewList_DataFrame):
        extra_rows = len(condition_macro_unique) - len(NewList_DataFrame)
        new_rows = pd.DataFrame(index=range(extra_rows), columns=NewList_DataFrame.columns)
        NewList_DataFrame = pd.concat([NewList_DataFrame, new_rows], ignore_index=True)
    NewList_DataFrame['ConditionMacro'] = condition_macro_unique
    if not os.path.exists(CONFIG_PATH):
        NewList_DataFrame.to_excel(CONFIG_PATH, sheet_name='List', index=False)
    else:
        with pd.ExcelWriter(CONFIG_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            NewList_DataFrame.to_excel(writer, sheet_name='List', index=False)


def reorder_sheets(input_file, output_file, desired_order):
    # Open the existing Excel file
    workbook = openpyxl.load_workbook(input_file)
    sheets = workbook.sheetnames
    print("Original sheet order:", sheets)

    # Create a new workbook
    new_workbook = openpyxl.Workbook()
    new_workbook.remove(new_workbook.active)  # Remove the default sheet

    # Define the font for all cells
    font = Font(name='Fira Code')
    alignment = Alignment(horizontal='left')

    # Add sheets in the desired order
    for sheet_name in desired_order:
        if sheet_name in sheets:
            sheet = workbook[sheet_name]
            new_sheet = new_workbook.create_sheet(sheet_name)
            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    new_cell.font = font
                    new_cell.alignment = alignment

    # Automatically adjust column widths
    for sheet_name in desired_order:
        new_sheet = new_workbook[sheet_name]
        for column in new_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            adjusted_width = (max_length + 5) * 1.2
            new_sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the modified Excel file
    new_workbook.save(output_file)
    print("Modified sheet order:", desired_order)


def main():
    WriteInputSignalToExcel()
    WriteConditionToExcel()
    WriteEVTToExcel()
    WriteTimeFlagToExcel()
    WriteConstMacroToExcel()
    WriteOutputInitializerToExcel()
    WriteActionToExcel()
    WriteListToExcel()
    reorder_sheets(CONFIG_PATH, CONFIG_PATH, order)


if __name__ == "__main__":
    main()
